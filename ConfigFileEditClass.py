from ConfigFileEditForm_ui  import Ui_ConfigFileEditForm
from PyQt5.QtWidgets import QApplication, QMainWindow, QTableView,QAbstractItemView, QHeaderView, QMenu, QMenuBar, QAction,QFileDialog,QMessageBox
from PyQt5.QtCore import Qt,QItemSelection, QItemSelectionModel, QModelIndex
from PyQt5.QtGui import QStandardItemModel, QStandardItem
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
import sys
import pandas as pd

#编辑Config文件
class ConfigFileEditClass(QtWidgets.QTableView, Ui_ConfigFileEditForm):
    def __init__(self, parent=None):
        super(ConfigFileEditClass, self).__init__(parent)
        self.setupUi(self)
        self.setWindowTitle("编辑Config文件")
        self.loadFile_Button.clicked.connect(self.readExcel)   
        self.Save_Button.clicked.connect(self.saveModifyExcel) 
        
    def appendRow(self):
        row = self.model.rowCount()
        itemlist = []     
        for col in range (self.cols):
            itemlist.append(QStandardItem('(row %s, column %s)'%(row, col)))          
        #添加一行
        self.model.appendRow(itemlist)
        #更新头信息
        self.model.setVerticalHeaderItem(row, QStandardItem('第{}行'.format(row+1)))
    
    def removeRow(self):
        #获取选中的行
        sel = self.tableView.selectionModel().selectedRows()
        print(sel)
        if sel:
            #下面删除时，选中多行中的最后一行，会被删掉；不选中，则默认第一行删掉
            index=self.tableView.currentIndex()
            print(index.row())
            self.model.removeRow(index.row())
            
    def onCurrentChanged(self,current, previous):
        #初始化时，previous.row() = -1，不显示信息
        if int(previous.row() < 0):
             return
        self.statusBar().showMessage('选中第{}行'.format(current.row()+1))
        
    
    #读取excel
    def readExcel(self):
        self.fileName=""
        #打开文件
        self.fileName, _ = QFileDialog.getOpenFileName(self, "打开文件", "", "Excel Files (*.xls *.xlsx)")
        if not self.fileName:
            return
        #读取文件
        workbook = openpyxl.load_workbook(self.fileName)
        #获取第一个sheet
        sheet = workbook.active
        #获取行数和列数
        rows = sheet.max_row
        cols = sheet.max_column
        #获取第一行作为表头
        headers = sheet[1]
        #创建model
        self.model = QStandardItemModel(rows, cols, self)
        
        insertrow=0
        #读取数据
        for row in sheet.iter_rows(min_row=2, max_row=rows, min_col=1, max_col=cols):
            for col, cell in enumerate(row):
                self.model.setItem(insertrow, col, QStandardItem(cell.value))
            insertrow+=1
        
        strlist=[]
        #设置表头
        for col, header in enumerate(headers):
            #self.model.setHeaderData(col, Qt.Horizontal, header)
            strlist.append(header.value)
        #填充数据

        self.model.setHorizontalHeaderLabels(strlist)
        '''
        for row in range(1, rows):
            items = sheet.merged_cells(row)
            for col, item in enumerate(items):
                self.model.setItem(row-1, col, QStandardItem(str(item))) 
        '''  
        
        #设置model
        self.Config_tableView.setModel(self.model)

        #设置列宽
        self.Config_tableView.setColumnWidth(0, 150)
        #设置行高
        self.Config_tableView.verticalHeader().setDefaultSectionSize(30)
        #设置选中模式
        self.Config_tableView.setSelectionBehavior(QAbstractItemView.SelectRows)
        #设置选中模式
        self.Config_tableView.setSelectionMode(QAbstractItemView.SingleSelection)
        #设置表格内容居中
        self.Config_tableView.horizontalHeader().setDefaultAlignment(Qt.AlignCenter)
        
    #保存修改后的数据到excel
    def saveModifyExcel(self):
        #获取文件名
        if not self.fileName:
            return
        #创建workbook
        workbook = openpyxl.load_workbook(self.fileName)
        #获取sheet
        sheet = workbook.active
        #获取mode行数和列数
        rows = self.model.rowCount()-1
        cols = self.model.columnCount()
        #写入数据
        for row  in range(0, rows):
            for col  in range(1,cols):
                itemtext=self.model.item(row, col).text()
                sheet.cell(row+2, col+1).value=itemtext
                     
        #保存文件
        workbook.save(self.fileName)